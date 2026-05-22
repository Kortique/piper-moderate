#!/usr/bin/env python3
"""
build_comparison_report.py
--------------------------
Generates lgbm_model_comparison.xlsx with 4 sheets:
  1. Модели — metadata for V6/V7/V8/V10
  2. Полный sweep — threshold sweep 0.30-0.60, all datasets
  3. Сводка — best THR per model per dataset
  4. V10 minor rule — V10 with/without minor>=0.72
"""
import json, re, ast, sqlite3, struct, numpy as np
import lightgbm as lgb
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent

# ── helpers ──────────────────────────────────────────────────────────────────
def sanitize(name):
    return name.replace(':', '_x').replace('"','').replace("'",'').replace('{','').replace('}','')

def build_X(items, feat_names):
    idx = {f: i for i, f in enumerate(feat_names)}
    X = np.zeros((len(items), len(feat_names)), dtype=np.float32)
    for i, it in enumerate(items):
        for k, v in it.get('underage_labels', {}).items():
            fk = sanitize(k)
            if fk in idx: X[i, idx[fk]] = float(v)
        for k, v in it.get('adult_labels', {}).items():
            fk = 'adult__' + sanitize(k)
            if fk in idx: X[i, idx[fk]] = float(v)
        for k, v in it.get('no_underage_labels', {}).items():
            fk = 'no_underage__' + sanitize(k)
            if fk in idx: X[i, idx[fk]] = float(v)
    return X

def score_all(items, model, feats, key):
    X = build_X(items, feats)
    preds = model.predict(X)
    for it, p in zip(items, preds):
        it[key] = float(p)

def eval_at(items, key, thr, minor_thr=None):
    child = [x for x in items if x['label'] == 'child']
    teen  = [x for x in items if x['label'] == 'teen']
    adult = [x for x in items if x['label'] == 'adult']
    if minor_thr:
        blocked = lambda x: x.get(key, 0) >= thr or x.get('minor', 0) >= minor_thr
    else:
        blocked = lambda x: x.get(key, 0) >= thr
    cr  = 100 * sum(1 for x in child if blocked(x)) / max(len(child), 1)
    tr  = 100 * sum(1 for x in teen  if blocked(x)) / max(len(teen),  1)
    fpr = 100 * sum(1 for x in adult if blocked(x)) / max(len(adult), 1)
    return cr, tr, fpr, len(child), len(teen), len(adult)

# ── Load models ───────────────────────────────────────────────────────────────
print('Loading models...')
js = (BASE / 'data' / 'lgbm_evaluate_v6.js').read_text()
v6_feats = json.loads(re.search(r'LGBM_FEATURES\s*=\s*(\[[\s\S]+?\]);', js).group(1))
v6_model = lgb.Booster(model_file=str(BASE / 'data' / 'lgbm_underage_v6.txt'))
v7_feats = json.loads((BASE / 'data' / 'lgbm_v7_features.json').read_text())
v7_model = lgb.Booster(model_file=str(BASE / 'data' / 'lgbm_underage_v7.txt'))
v8_feats = json.loads((BASE / 'data' / 'lgbm_v8_features.json').read_text())
v8_model = lgb.Booster(model_file=str(BASE / 'data' / 'lgbm_underage_v8.txt'))
v10_feats = json.loads((BASE / 'data' / 'lgbm_v10_features.json').read_text())
v10_model = lgb.Booster(model_file=str(BASE / 'data' / 'lgbm_underage_v10.txt'))
print(f'  V6:{len(v6_feats)}f  V7:{len(v7_feats)}f  V8:{len(v8_feats)}f  V10:{len(v10_feats)}f')

# ── Load LS items (rescored + fallback) ───────────────────────────────────────
print('Loading LS holdout...')
ls_rescored = {}
for r in json.loads((BASE / 'data' / 'ls_holdout_rescored.json').read_text()):
    if r.get('done') and r.get('label') in ('child', 'teen', 'adult'):
        ls_rescored[r['task_id']] = r

raw = (BASE / 'qwen3_age_results.json').read_bytes().rstrip(b'\x00').decode('utf-8')
ls_orig_data = json.loads(raw)
ls_items = []
for v in ls_orig_data.values():
    lbl = v.get('category')
    if lbl not in ('child', 'teen', 'adult'):
        continue
    tid = v['task_id']
    if tid in ls_rescored:
        r = ls_rescored[tid]
        ls_items.append({'id': f'ls_{tid}', 'label': lbl, 'minor': r.get('minor', 0),
                         'source': 'LS', 'rescored': True,
                         'underage_labels': r.get('underage_labels', {}),
                         'adult_labels': r.get('adult_labels', {}),
                         'no_underage_labels': r.get('no_underage_labels', {})})
    else:
        siglip_raw = v.get('siglip2_details')
        siglip2 = ast.literal_eval(siglip_raw) if isinstance(siglip_raw, str) else (siglip_raw or {})
        det = siglip2.get('underage', {})
        lbls = det.get('labels', {})
        ls_items.append({'id': f'ls_{tid}', 'label': lbl, 'minor': det.get('minor', 0),
                         'source': 'LS', 'rescored': False,
                         'underage_labels': lbls.get('underage') or {},
                         'adult_labels': lbls.get('adult') or {},
                         'no_underage_labels': {}})
print(f'  {len(ls_items)} LS items ({sum(1 for x in ls_items if x["rescored"])} rescored)')

# ── Load Grafana pool ─────────────────────────────────────────────────────────
print('Loading Grafana pool...')
cands = sorted((BASE / 'backups').glob('gallery_*.db'), reverse=True)
data = bytearray(cands[0].read_bytes())
struct.pack_into('>I', data, 28, len(data) // 4096)
tmp = Path('/tmp/_eval_all.db')
tmp.write_bytes(bytes(data))
conn = sqlite3.connect(str(tmp))
conn.row_factory = sqlite3.Row
rows = conn.execute("""
    SELECT id, label, piper_result, export_batch
    FROM grafana_pool
    WHERE (deleted IS NULL OR deleted=0)
    AND label IN ('child','teen','adult')
    AND piper_result IS NOT NULL
""").fetchall()
conn.close()

v9_317 = {}
p317 = BASE / 'data' / 'v9_317_scores.json'
if p317.exists():
    for r in json.loads(p317.read_text()):
        if r.get('done'):
            v9_317[r['id']] = r

grafana_items = []
for r in rows:
    gid, lbl, eb = r['id'], r['label'], r['export_batch']
    is_317 = (eb == '2026-05-20 UTC')
    if gid in v9_317:
        v = v9_317[gid]
        grafana_items.append({'id': gid, 'label': lbl, 'minor': v.get('minor', 0),
                              'source': '317-session' if is_317 else 'Grafana',
                              'rescored': True,
                              'underage_labels': v.get('underage_labels', {}),
                              'adult_labels': v.get('adult_labels', {}),
                              'no_underage_labels': v.get('no_underage_labels', {})})
    else:
        try:
            pr = json.loads(r['piper_result'])
            det = (pr.get('siglip2_details') or {}).get('underage', {})
            lbls = det.get('labels', {})
            grafana_items.append({'id': gid, 'label': lbl, 'minor': det.get('minor', 0),
                                  'source': '317-session' if is_317 else 'Grafana',
                                  'rescored': False,
                                  'underage_labels': lbls.get('underage') or {},
                                  'adult_labels': lbls.get('adult') or {},
                                  'no_underage_labels': {}})
        except:
            pass
print(f'  {len(grafana_items)} Grafana items')

# Merge & deduplicate
all_items = []
seen = set()
for it in ls_items + grafana_items:
    if it['id'] not in seen:
        seen.add(it['id'])
        all_items.append(it)

by_src = {}
for it in all_items:
    by_src[it['source']] = by_src.get(it['source'], 0) + 1
print(f'Total: {len(all_items)}  {by_src}')

# ── Score all models ──────────────────────────────────────────────────────────
print('Scoring...')
score_all(all_items, v6_model, v6_feats, 'v6')
score_all(all_items, v7_model, v7_feats, 'v7')
score_all(all_items, v8_model, v8_feats, 'v8')
score_all(all_items, v10_model, v10_feats, 'v10')
print('Done scoring.')

# ── Excel styles ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', start_color='1F3864')
SUBHDR_FILL = PatternFill('solid', start_color='2E75B6')
ALT_FILL    = PatternFill('solid', start_color='EEF3FB')
GREEN_FILL  = PatternFill('solid', start_color='C6EFCE')
ORANGE_FILL = PatternFill('solid', start_color='FFEB9C')
RED_FILL    = PatternFill('solid', start_color='FFC7CE')
WHITE_FILL  = PatternFill('solid', start_color='FFFFFF')
HDR_FONT    = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SUBHDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
CENTER      = Alignment(horizontal='center', vertical='center')
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hcell(ws, row, col, val, fill=HEADER_FILL, font=HDR_FONT):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill; c.font = font; c.alignment = CENTER; c.border = BORDER
    return c

def dcell(ws, row, col, val, fill=WHITE_FILL, bold=False, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill
    c.font = Font(name='Arial', size=10, bold=bold)
    c.alignment = CENTER; c.border = BORDER
    if fmt: c.number_format = fmt
    return c

THRESHOLDS = [round(t, 2) for t in np.arange(0.60, 0.25, -0.05)]
MODELS = [
    ('V6',  'v6',  312, 'siglip5 (underage+adult)',          713, 0.80, '2026-05-20'),
    ('V7',  'v7',  314, 'siglip5 +317-session',              713, 0.45, '2026-05-20'),
    ('V8',  'v8',  155, 'siglip5 +317-session (more feats)', 713, 0.45, '2026-05-20'),
    ('V10', 'v10', 510, '867 tags incl. 137 no_underage__*', 867, 0.45, '2026-05-21'),
]
# MODELS tuple: (name, key, n_feats, tagset_desc, n_tags_total, default_thr, train_date)
DATASETS = [
    ('ALL (3120)',         lambda x: True),
    ('LS (rescored 2212)', lambda x: x['source'] == 'LS' and x['rescored']),
    ('317-session (317)', lambda x: x['source'] == '317-session'),
    ('Grafana non-317 (591)', lambda x: x['source'] == 'Grafana'),
]

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: METADATA
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Модели'

ws1.merge_cells('A1:H1')
c = ws1['A1']
c.value = 'Сравнение моделей LGBM — метаданные'
c.fill = HEADER_FILL; c.font = Font(name='Arial', bold=True, color='FFFFFF', size=13)
c.alignment = CENTER; c.border = BORDER

for ci, h in enumerate(['Модель','Фич','Тег-сет','Тегов всего','Default THR','Обучен','Правило блокировки','Файл модели'], 1):
    hcell(ws1, 2, ci, h, fill=SUBHDR_FILL, font=SUBHDR_FONT)

meta = [
    ('V6',  312, 'siglip5: underage + adult',             713, 0.80, '2026-05-20', 'lgbm ≥ 0.80', 'lgbm_underage_v6.txt'),
    ('V7',  314, 'siglip5 + 317-session hard examples',   713, 0.45, '2026-05-20', 'lgbm ≥ 0.45', 'lgbm_underage_v7.txt'),
    ('V8',  355, 'siglip5 + 317-session hard examples',   713, 0.45, '2026-05-20', 'lgbm ≥ 0.45', 'lgbm_underage_v8.txt'),
    ('V10', 510, '867 tags: siglip5 + 160 no_underage__*',867, 0.45, '2026-05-21', 'lgbm ≥ 0.45', 'lgbm_underage_v10.txt'),
]
notes = {
    'V6':  'Deployed in production. Оценка через JS (lgbm_evaluate_v6.js). Обучен на LS+grafana (713 тегов).',
    'V7':  '317-session hard examples × weight 3.0. LS holdout с siglip5 (no_underage=0).',
    'V8':  'Как V7, но больше фич (+41). Те же данные, та же структура.',
    'V10': 'LS holdout rescored через ce79f7e299 (867 тегов). Устраняет distribution mismatch V7/V8/V9. 137 no_underage__ фич.',
}
for ri, row in enumerate(meta, 3):
    fill = WHITE_FILL if ri % 2 else ALT_FILL
    for ci, val in enumerate(row, 1):
        c = dcell(ws1, ri, ci, val, fill=fill)
        if ci == 1: c.font = Font(name='Arial', bold=True, size=10)
        if ci == 5: c.number_format = '0.00'
    note_c = ws1.cell(row=ri, column=9, value=notes[row[0]])
    note_c.font = Font(name='Arial', size=9, italic=True, color='595959')
    note_c.alignment = Alignment(horizontal='left', wrap_text=True)
    note_c.border = BORDER

# Dataset block
ws1.cell(row=8, column=1).value = ''
hcell(ws1, 8, 1, 'Датасет', fill=SUBHDR_FILL, font=SUBHDR_FONT)
hcell(ws1, 8, 2, 'Кол-во', fill=SUBHDR_FILL, font=SUBHDR_FONT)
ws1.merge_cells('C8:I8')
hcell(ws1, 8, 3, 'Описание', fill=SUBHDR_FILL, font=SUBHDR_FONT)

ds_info = [
    ('Label Studio (LS)', 2212, 'Rescored через Piper ce79f7e299 (867 тегов). 2187 done + 25 fallback на siglip5.'),
    ('Grafana 317-session', 317, 'Hard examples: AI-генерированный взрослый контент. Rescored с 867 тегами. export_batch=2026-05-20 UTC.'),
    ('Grafana non-317', 591, 'Старые items из grafana_pool. Оценены siglip5. no_underage=0 (до rescore).'),
    ('ИТОГО', 3120, ''),
]
for ri, (name, cnt, desc) in enumerate(ds_info, 9):
    fill = ALT_FILL if ri % 2 else WHITE_FILL
    is_total = (name == 'ИТОГО')
    dcell(ws1, ri, 1, name, fill=fill, bold=is_total)
    dcell(ws1, ri, 2, cnt, fill=fill, bold=is_total)
    ws1.merge_cells(f'C{ri}:I{ri}')
    c = ws1.cell(row=ri, column=3, value=desc)
    c.font = Font(name='Arial', size=10, bold=is_total)
    c.border = BORDER; c.alignment = Alignment(horizontal='left')

for col_letter, width in zip('ABCDEFGHI', [12,8,40,12,11,12,18,30,50]):
    ws1.column_dimensions[col_letter].width = width

print('Sheet 1 done')

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: FULL SWEEP
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Полный sweep')

# Row 1: model names spanning 3 cols each
ws2.merge_cells('A1:A3')
c = ws2['A1']; c.value='Датасет'; c.fill=HEADER_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER
ws2.merge_cells('B1:B3')
c = ws2['B1']; c.value='THR'; c.fill=HEADER_FILL; c.font=HDR_FONT; c.alignment=CENTER; c.border=BORDER

col = 3
model_cols = {}
for mname, mkey, mfeats, mtags, mdef_thr, mdate in [(m[0],m[1],m[2],m[3],m[4],m[6]) for m in MODELS]:
    ws2.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
    c = ws2.cell(row=1, column=col, value=f'{mname}  ({mfeats}f)')
    c.fill=SUBHDR_FILL; c.font=SUBHDR_FONT; c.alignment=CENTER; c.border=BORDER
    ws2.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+2)
    c = ws2.cell(row=2, column=col, value=f'Default THR: {mdef_thr}')
    c.fill=HEADER_FILL; c.font=Font(name='Arial',bold=False,color='FFFFFF',size=9,italic=True)
    c.alignment=CENTER; c.border=BORDER
    for ci, lbl in enumerate(['Child%','Teen%','FPR%']):
        hcell(ws2, 3, col+ci, lbl, fill=SUBHDR_FILL, font=SUBHDR_FONT)
    model_cols[mkey] = col
    col += 3

row = 4
for ds_name, ds_filter in DATASETS:
    ds_items = [x for x in all_items if ds_filter(x)]
    cn = sum(1 for x in ds_items if x['label']=='child')
    tn = sum(1 for x in ds_items if x['label']=='teen')
    an = sum(1 for x in ds_items if x['label']=='adult')
    total_cols = 2 + 3 * len(MODELS)
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    c = ws2.cell(row=row, column=1,
                 value=f'{ds_name}   child={cn}, teen={tn}, adult={an}, total={len(ds_items)}')
    c.fill=HEADER_FILL; c.font=HDR_FONT
    c.alignment=Alignment(horizontal='left',vertical='center'); c.border=BORDER
    row += 1

    for ti, thr in enumerate(THRESHOLDS):
        fill = ALT_FILL if ti % 2 else WHITE_FILL
        dcell(ws2, row, 1, ds_name, fill=fill)
        dcell(ws2, row, 2, thr, fill=fill, fmt='0.00')
        for mname, mkey, mfeats2, mtags2, mdef_thr2, ntags2, mdate2 in MODELS:
            cr, tr, fpr, cn2, tn2, an2 = eval_at(ds_items, mkey, thr)
            sc = model_cols[mkey]
            # child
            cf = GREEN_FILL if cr>=95 else (ORANGE_FILL if cr>=90 else RED_FILL)
            c = ws2.cell(row=row, column=sc, value=round(cr,1))
            c.fill=cf; c.font=Font(name='Arial',bold=(cr>=95),size=10); c.alignment=CENTER; c.border=BORDER; c.number_format='0.0'
            # teen
            tf = GREEN_FILL if tr>=85 else (ORANGE_FILL if tr>=70 else fill)
            c = ws2.cell(row=row, column=sc+1, value=round(tr,1))
            c.fill=tf; c.font=Font(name='Arial',size=10); c.alignment=CENTER; c.border=BORDER; c.number_format='0.0'
            # fpr
            ff = GREEN_FILL if fpr<=5 else (ORANGE_FILL if fpr<=15 else RED_FILL)
            c = ws2.cell(row=row, column=sc+2, value=round(fpr,1))
            c.fill=ff; c.font=Font(name='Arial',size=10); c.alignment=CENTER; c.border=BORDER; c.number_format='0.0'
        row += 1
    row += 1

ws2.column_dimensions['A'].width = 24
ws2.column_dimensions['B'].width = 6
for i in range(3, col):
    ws2.column_dimensions[get_column_letter(i)].width = 9
ws2.freeze_panes = 'C4'
print('Sheet 2 done')

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: SUMMARY
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Сводка')
ws3.merge_cells('A1:J1')
c = ws3['A1']
c.value = 'Лучший порог: child ≥ 95%, минимальный FPR  (LGBM-only, без minor rule)'
c.fill=HEADER_FILL; c.font=Font(name='Arial',bold=True,color='FFFFFF',size=12); c.alignment=CENTER; c.border=BORDER

for ci, h in enumerate(['Датасет','Модель','Best THR','Child%','Teen%','FPR%','Child n','Teen n','Adult n','Примечание'], 1):
    hcell(ws3, 2, ci, h, fill=SUBHDR_FILL, font=SUBHDR_FONT)

row3 = 3
for ds_name, ds_filter in DATASETS:
    ds_items = [x for x in all_items if ds_filter(x)]
    for mname, mkey, mfeats, mtags, mdef_thr, mdate in [(m[0],m[1],m[2],m[3],m[4],m[6]) for m in MODELS]:
        best = None
        for thr in THRESHOLDS:
            cr, tr, fpr, cn, tn, an = eval_at(ds_items, mkey, thr)
            if cr >= 95.0:
                if best is None or fpr < best[3]:
                    best = (thr, cr, tr, fpr, cn, tn, an)
        flag = '✓'
        if best is None:
            for thr in THRESHOLDS:
                cr, tr, fpr, cn, tn, an = eval_at(ds_items, mkey, thr)
                if best is None or cr > best[1]:
                    best = (thr, cr, tr, fpr, cn, tn, an)
            flag = '✗ no 95%'

        fill = ALT_FILL if row3 % 2 == 0 else WHITE_FILL
        thr_str = f'{best[0]:.2f} {flag}'
        note = '≥95% ✓' if flag == '✓' else 'Не достигает 95%'
        vals = [ds_name, mname, thr_str, round(best[1],1), round(best[2],1),
                round(best[3],1), best[4], best[5], best[6], note]
        for ci, val in enumerate(vals, 1):
            c = dcell(ws3, row3, ci, val, fill=fill)
            if ci == 2: c.font = Font(name='Arial', bold=True, size=10)
            if ci == 4:
                c.fill = GREEN_FILL if best[1]>=95 else ORANGE_FILL
            if ci == 6:
                c.fill = GREEN_FILL if best[3]<=5 else (ORANGE_FILL if best[3]<=15 else RED_FILL)
        row3 += 1
    row3 += 1

for col_letter, width in zip('ABCDEFGHIJ', [26,8,13,9,9,9,9,9,9,16]):
    ws3.column_dimensions[col_letter].width = width
print('Sheet 3 done')

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: V10 with/without minor>=0.72
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('V10 minor rule')
ws4.merge_cells('A1:I1')
c = ws4['A1']
c.value = 'V10: LGBM-only vs LGBM OR minor≥0.72'
c.fill=HEADER_FILL; c.font=Font(name='Arial',bold=True,color='FFFFFF',size=12); c.alignment=CENTER; c.border=BORDER

for ci, h in enumerate(['Датасет','THR','Child% LGBM','Teen% LGBM','FPR% LGBM','','Child% +minor','Teen% +minor','FPR% +minor'], 1):
    hcell(ws4, 2, ci, h, fill=SUBHDR_FILL if h else HEADER_FILL, font=SUBHDR_FONT)

row4 = 3
for ds_name, ds_filter in DATASETS:
    ds_items = [x for x in all_items if ds_filter(x)]
    child = [x for x in ds_items if x['label']=='child']
    teen  = [x for x in ds_items if x['label']=='teen']
    adult = [x for x in ds_items if x['label']=='adult']

    ws4.merge_cells(start_row=row4, start_column=1, end_row=row4, end_column=9)
    c = ws4.cell(row=row4, column=1,
                 value=f'{ds_name}   child={len(child)}, teen={len(teen)}, adult={len(adult)}')
    c.fill=HEADER_FILL; c.font=HDR_FONT
    c.alignment=Alignment(horizontal='left',vertical='center'); c.border=BORDER
    row4 += 1

    for ti, thr in enumerate(THRESHOLDS):
        fill = ALT_FILL if ti % 2 else WHITE_FILL
        cr1 = 100*sum(1 for x in child if x.get('v10',0)>=thr)/max(len(child),1)
        tr1 = 100*sum(1 for x in teen  if x.get('v10',0)>=thr)/max(len(teen),1)
        fp1 = 100*sum(1 for x in adult if x.get('v10',0)>=thr)/max(len(adult),1)
        cr2 = 100*sum(1 for x in child if x.get('v10',0)>=thr or x.get('minor',0)>=0.72)/max(len(child),1)
        tr2 = 100*sum(1 for x in teen  if x.get('v10',0)>=thr or x.get('minor',0)>=0.72)/max(len(teen),1)
        fp2 = 100*sum(1 for x in adult if x.get('v10',0)>=thr or x.get('minor',0)>=0.72)/max(len(adult),1)

        dcell(ws4, row4, 1, ds_name, fill=fill)
        dcell(ws4, row4, 2, thr, fill=fill, fmt='0.00')
        for ci, val in enumerate([cr1, tr1, fp1, None, cr2, tr2, fp2], 3):
            if val is None:
                ws4.cell(row=row4, column=ci).fill = HEADER_FILL
                continue
            orig_ci = ci
            c = dcell(ws4, row4, ci, round(val,1), fill=fill, fmt='0.0')
            if orig_ci in (3, 7):  # child
                c.fill = GREEN_FILL if val>=95 else (ORANGE_FILL if val>=90 else RED_FILL)
                if val >= 95: c.font = Font(name='Arial', bold=True, size=10)
            elif orig_ci in (5, 9):  # fpr
                c.fill = GREEN_FILL if val<=5 else (ORANGE_FILL if val<=15 else RED_FILL)
        row4 += 1
    row4 += 1

for col_letter, width in zip('ABCDEFGHI', [24,6,13,13,12,3,13,13,12]):
    ws4.column_dimensions[col_letter].width = width
print('Sheet 4 done')

# ── Save ──────────────────────────────────────────────────────────────────────
OUT = BASE / 'lgbm_model_comparison.xlsx'
wb.save(str(OUT))
print(f'\nSaved → {OUT}')
